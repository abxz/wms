"""导入导出模块"""
from fastapi import APIRouter, UploadFile, File, Query, HTTPException
from core.database import all_, add
from core.utils import generate_id
from modules.products import service as prod_svc
from modules.employees import service as emp_svc
from modules.inventory import service as inv_svc
from modules.suppliers import service as sup_svc
from modules.warehouses import service as wh_svc

router = APIRouter(prefix="/api/import", tags=["导入导出"])

def register(app):
    app.include_router(router)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
ALLOWED_MIME = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
]

def _validate_excel(file: UploadFile) -> bool:
    if file.size and file.size > MAX_FILE_SIZE:
        raise HTTPException(413, "文件超过 10MB 限制")
    if file.content_type not in ALLOWED_MIME:
        raise HTTPException(415, "仅支持 .xlsx 文件")
    return True

@router.post("/main-data")
async def import_main_data(file: UploadFile = File(...)):
    """模板A：主数据（仓库+商品+供应商+库存+发票号码）"""
    _validate_excel(file)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")

    headers = ["仓库名称","仓库地址","商品名称","商品规格","商品条码",
               "库存预警","供应商名称","库存数量","发票号码","单位","单价"]
    
    results = {"total": len(rows), "success": 0, "errors": []}
    for i, row in enumerate(rows):
        line_no = i + 2
        try:
            wh_name, wh_addr, p_name, p_spec, p_barcode, min_stk, sup_name, inv_qty, inv_no, unit, price = (
                (str(v).strip() if v else "") for v in (row + (None,)*11)[:11]
            )
            if not wh_name or not p_name:
                results["errors"].append(f"第{line_no}行：仓库名称或商品名称为空")
                continue

            # 查找或创建仓库
            whs = wh_svc.get_all_warehouses()
            wh = next((w for w in whs if w["name"] == wh_name), None)
            if not wh:
                wh = wh_svc.create_warehouse({"name": wh_name, "address": wh_addr})
            # 查找或创建供应商
            sups = sup_svc.get_all_suppliers()
            sup = next((s for s in sups if s["name"] == sup_name), None) if sup_name else None
            if not sup and sup_name:
                sup = sup_svc.create_supplier({"name": sup_name, "contact": "", "phone": ""})
            # 创建商品
            prod = prod_svc.create_product({
                "name": p_name, "spec": p_spec, "barcode": p_barcode,
                "warehouse_id": wh["id"], "unit": unit or "个",
                "supplier_id": sup["id"] if sup else "",
                "min_stock": float(min_stk) if min_stk else 0, "price": float(price) if price else 0,
                "invoice_number": inv_no
            })
            # 初始化库存
            if inv_qty:
                try:
                    inv_svc.update_stock(prod["id"], float(inv_qty))
                except:
                    pass
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"第{line_no}行：{str(e)}")
    return results


@router.post("/employees")
async def import_employees(file: UploadFile = File(...)):
    """模板B：员工（完整字段）"""
    _validate_excel(file)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")

    results = {"total": len(rows), "success": 0, "errors": []}
    for i, row in enumerate(rows):
        line_no = i + 2
        try:
            name, dept, position, job_type, role, education, id_card, address, quota = (
                (str(v).strip() if v else "") for v in (row + (None,)*9)[:9]
            )
            if not name:
                results["errors"].append(f"第{line_no}行：姓名为空")
                continue
            role = role if role in ("super_admin", "admin", "claimer") else "claimer"
            emp_svc.create_employee({
                "name": name, "department": dept, "position": position,
                "job_type": job_type, "role": role, "education": education,
                "id_card": id_card, "address": address,
                "monthly_quota": float(quota) if quota else 1000,
            })
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"第{line_no}行：{str(e)}")
    return results


@router.post("/orders")
async def import_orders(file: UploadFile = File(...)):
    """模板C：出入库（按日期排序，先入库后出库）"""
    _validate_excel(file)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")

    results = {"total": len(rows), "success": 0, "errors": []}
    # 先按日期排序，入库优先
    parsed = []
    for i, row in enumerate(rows):
        typ, order_no, p_name, qty_str, operator, claimer, location, date_str = (
            (str(v).strip() if v else "") for v in (row + (None,)*8)[:8]
        )
        if typ not in ("入库", "出库"):
            results["errors"].append(f"第{i+2}行：类型必须为'入库'或'出库'")
            continue
        parsed.append({
            "line": i+2, "type": "inbound" if typ == "入库" else "outbound",
            "order_no": order_no, "product_name": p_name, "qty": qty_str,
            "operator": operator, "claimer": claimer, "location": location, "date": date_str
        })
    # 按日期排序，同天入库优先
    parsed.sort(key=lambda r: (r["date"] or "0") + ("0" if r["type"]=="inbound" else "1"))
    
    from modules.inbound import service as in_svc
    from modules.outbound import service as out_svc
    from modules.inventory import service as inv_svc

    for r in parsed:
        try:
            qty = float(r["qty"]) if r["qty"] else 0
            if qty <= 0:
                results["errors"].append(f"第{r['line']}行：数量必须大于0")
                continue
            # 查找商品
            products = prod_svc.get_all_products()
            prod = next((p for p in products if p["name"] == r["product_name"]), None)
            if not prod:
                results["errors"].append(f"第{r['line']}行：商品 '{r['product_name']}' 不存在")
                continue
            # 查找操作人
            emps = emp_svc.list_employees(search=r["operator"])
            emp = next((e for e in emps.get("items",[]) if r["operator"] in e.get("name","")), None)

            if r["type"] == "inbound":
                in_svc.create_inbound({
                    "items": [{"product_id": prod["id"], "quantity": qty, "price": 0}],
                    "supplier_name": "", "admin_name": r["operator"]
                })
            else:  # outbound
                # 检查库存
                inv = inv_svc.get_stock_by_product(prod["id"])
                cur_qty = inv.get("quantity", 0) if inv else 0
                if cur_qty < qty:
                    results["errors"].append(
                        f"第{r['line']}行：'{r['product_name']}' 库存不足（当前{cur_qty}，需要{qty}）")
                    continue
                out_svc.create_outbound({
                    "items": [{"product_id": prod["id"], "quantity": qty, "price": 0}],
                    "claimer_name": r["claimer"], "admin_name": r["operator"],
                    "usage_location": r["location"]
                })
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"第{r['line']}行：{str(e)}")
    return results


@router.get("/template/main-data")
def download_main_data_template():
    """下载主数据导入模板"""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主数据导入模板"
    headers = ["仓库名称","仓库地址","商品名称","商品规格","商品条码",
               "库存预警数","供应商名称","库存数量","发票号码","单位","单价"]
    ws.append(headers)
    ws.append(["主仓库","","示例轴承","6205ZZ","6901234567890","10","ABC供应商","100","INV-2026-0001","个","25"])
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=main_data_template.xlsx"})


@router.get("/template/employees")
def download_employee_template():
    """下载员工导入模板（完整字段）"""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工导入模板"
    headers = ["姓名","部门","岗位","工种","角色","学历","身份证号","地址","月度额度"]
    ws.append(headers)
    ws.append(["张三","生产部","班组长","爆破工","claimer","大专","410102199001010001","河南省郑州市","1000"])
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=employee_template.xlsx"})


@router.get("/template/orders")
def download_order_template():
    """下载出入库导入模板"""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出入库导入模板"
    headers = ["类型","单号","产品名称","数量","操作人","领料人","使用地点","日期"]
    ws.append(headers)
    ws.append(["入库","","轴承","100","王主管","","","2026-05-14"])
    ws.append(["出库","","轴承","10","王主管","张三","3号车间","2026-05-14"])
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=order_template.xlsx"})


# ═══════════════════════════════════════════════════════════════
# 导出
# ═══════════════════════════════════════════════════════════════

def _excel_response(wb, filename: str):
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/export/products")
def export_products():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "商品"
    ws.append(["ID","SKU","名称","规格","分类","单位","单价","条码","供应商ID","仓库ID","最低库存","最高库存","备注"])
    for p in prod_svc.get_all_products():
        ws.append([p.get(k,"") for k in ["id","sku","name","spec","category","unit","price","barcode","supplier_id","warehouse_id","min_stock","max_stock","remark"]])
    return _excel_response(wb, "products.xlsx")


@router.get("/export/suppliers")
def export_suppliers():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "供应商"
    ws.append(["ID","名称","联系人","电话","地址","备注"])
    for s in sup_svc.get_all_suppliers():
        ws.append([s.get(k,"") for k in ["id","name","contact","phone","address","remark"]])
    return _excel_response(wb, "suppliers.xlsx")


@router.get("/export/employees")
def export_employees():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "员工"
    ws.append(["ID","姓名","工号","部门","岗位","工种","角色","学历","身份证号","地址","月度额度","已用额度","状态"])
    for e in emp_svc.list_employees(size=9999).get("items", []):
        ws.append([e.get(k,"") for k in ["id","name","employee_no","department","position","job_type","role","education","id_card","address","monthly_quota","monthly_used","active"]])
    return _excel_response(wb, "employees.xlsx")


@router.get("/export/inventory")
def export_inventory():
    import openpyxl
    from core.database import all_
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "库存"
    ws.append(["商品ID","商品名称","库位ID","数量","更新时间"])
    products = {p["id"]: p["name"] for p in prod_svc.get_all_products()}
    for inv in all_("inventory"):
        ws.append([inv.get("product_id",""), products.get(inv.get("product_id",""),""),
                   inv.get("location_id",""), inv.get("quantity",0), inv.get("updated_at","")])
    return _excel_response(wb, "inventory.xlsx")

# ═══════════════════════════════════════════════════════════════
# master_data 导入导出
# ═══════════════════════════════════════════════════════════════

@router.get("/export/master-products")
def export_master_products():
    import openpyxl
    from core.database import all_
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础商品"
    ws.append(["ID","名称","SKU","规格","单位","分类","条码","单价","最低库存","最高库存","供应商ID","仓库ID","库位ID","备注"])
    for p in all_("master_products"):
        ws.append([p.get(k,"") for k in ["id","name","sku","spec","unit","category","barcode","price","min_stock","max_stock","supplier_id","warehouse_id","location_id","remark"]])
    return _excel_response(wb, "master_products.xlsx")

@router.get("/export/master-suppliers")
def export_master_suppliers():
    import openpyxl
    from core.database import all_
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础供应商"
    ws.append(["ID","名称","联系人","电话","地址","备注"])
    for s in all_("master_suppliers"):
        ws.append([s.get(k,"") for k in ["id","name","contact","phone","address","remark"]])
    return _excel_response(wb, "master_suppliers.xlsx")

@router.get("/export/master-employees")
def export_master_employees():
    import openpyxl
    from core.database import all_
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础员工"
    ws.append(["ID","姓名","工号","部门","备注"])
    for e in all_("master_employees"):
        ws.append([e.get(k,"") for k in ["id","name","employee_no","department","remark"]])
    return _excel_response(wb, "master_employees.xlsx")


@router.post("/import/master-products")
async def import_master_products(file: UploadFile = File(...)):
    _validate_excel(file)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")
    from modules.master_data import service as md_svc
    results = {"total": len(rows), "success": 0, "errors": []}
    for i, row in enumerate(rows):
        line_no = i + 2
        try:
            name = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
            if not name:
                results["errors"].append(f"第{line_no}行：名称为空")
                continue
            data = {
                "name": name,
                "sku": str(row[2] or "").strip() if len(row) > 2 else "",
                "spec": str(row[3] or "").strip() if len(row) > 3 else "",
                "unit": str(row[4] or "个").strip() if len(row) > 4 else "个",
                "category": str(row[5] or "").strip() if len(row) > 5 else "",
                "barcode": str(row[6] or "").strip() if len(row) > 6 else "",
                "price": float(row[7]) if len(row) > 7 and row[7] else 0,
                "min_stock": float(row[8]) if len(row) > 8 and row[8] else 0,
                "supplier_id": str(row[10] or "").strip() if len(row) > 10 else "",
                "warehouse_id": str(row[11] or "").strip() if len(row) > 11 else "",
                "location_id": str(row[12] or "").strip() if len(row) > 12 else "",
                "remark": str(row[13] or "").strip() if len(row) > 13 else "",
            }
            md_svc.create_product(data)
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"第{line_no}行：{str(e)}")
    return results


@router.post("/import/master-suppliers")
async def import_master_suppliers(file: UploadFile = File(...)):
    _validate_excel(file)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")
    from modules.master_data import service as md_svc
    results = {"total": len(rows), "success": 0, "errors": []}
    for i, row in enumerate(rows):
        line_no = i + 2
        try:
            name = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
            if not name:
                results["errors"].append(f"第{line_no}行：名称为空")
                continue
            data = {
                "name": name,
                "contact": str(row[2] or "").strip() if len(row) > 2 else "",
                "phone": str(row[3] or "").strip() if len(row) > 3 else "",
                "address": str(row[4] or "").strip() if len(row) > 4 else "",
                "remark": str(row[5] or "").strip() if len(row) > 5 else "",
            }
            md_svc.create_supplier(data)
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"第{line_no}行：{str(e)}")
    return results


@router.post("/import/master-employees")
async def import_master_employees(file: UploadFile = File(...)):
    _validate_excel(file)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")
    from modules.master_data import service as md_svc
    results = {"total": len(rows), "success": 0, "errors": []}
    for i, row in enumerate(rows):
        line_no = i + 2
        try:
            name = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
            if not name:
                results["errors"].append(f"第{line_no}行：姓名为空")
                continue
            data = {
                "name": name,
                "employee_no": str(row[2] or "").strip() if len(row) > 2 else "",
                "department": str(row[3] or "").strip() if len(row) > 3 else "",
                "remark": str(row[4] or "").strip() if len(row) > 4 else "",
            }
            md_svc.create_employee(data)
            results["success"] += 1
        except Exception as e:
            results["errors"].append(f"第{line_no}行：{str(e)}")
    return results


@router.get("/template/master-products")
def download_master_products_template():
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础商品导入模板"
    headers = ["ID(留空新增)","名称","SKU","规格","单位","分类","条码","单价","最低库存","供应商ID","仓库ID","库位ID","备注"]
    ws.append(headers)
    ws.append(["","示例水泥","CE-001","P.O 42.5","吨","水泥","6901234567890","450","10","","","",""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=master_products_template.xlsx"})


@router.get("/template/master-suppliers")
def download_master_suppliers_template():
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础供应商导入模板"
    headers = ["ID(留空新增)","名称","联系人","电话","地址","备注"]
    ws.append(headers)
    ws.append(["","示例供应商","张三","13800138000","北京市",""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=master_suppliers_template.xlsx"})


@router.get("/template/master-employees")
def download_master_employees_template():
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "基础员工导入模板"
    headers = ["ID(留空新增)","姓名","工号","部门","备注"]
    ws.append(headers)
    ws.append(["","张三","EMP-001","机加工车间",""])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=master_employees_template.xlsx"})


# ═══════════════════════════════════════════════════════════════
# master_config 配置项导入导出（部门/岗位/工种/角色）
# ═══════════════════════════════════════════════════════════════

CONFIG_TYPE_LABELS = {
    "departments": "部门",
    "positions": "岗位",
    "job_types": "工种",
    "roles": "角色",
}


@router.get("/export/config/{config_type}")
def export_config(config_type: str):
    """导出配置项（部门/岗位/工种/角色）"""
    if config_type not in CONFIG_TYPE_LABELS:
        raise HTTPException(400, f"不支持的配置类型: {config_type}")
    import openpyxl
    import json
    from core.database import get_by
    label = CONFIG_TYPE_LABELS[config_type]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{label}列表"
    ws.append([label])
    row = get_by("system_config", "key", config_type)
    items = json.loads(row.get("value", "[]")) if row else []
    for item in items:
        ws.append([item])
    return _excel_response(wb, f"{config_type}.xlsx")


@router.get("/template/config/{config_type}")
def download_config_template(config_type: str):
    """下载配置项导入模板（部门/岗位/工种/角色）"""
    if config_type not in CONFIG_TYPE_LABELS:
        raise HTTPException(400, f"不支持的配置类型: {config_type}")
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    label = CONFIG_TYPE_LABELS[config_type]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{label}导入模板"
    ws.append([label])
    ws.append([f"示例{label}1"])
    ws.append([f"示例{label}2"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={config_type}_template.xlsx"})


def _levenshtein(s1: str, s2: str) -> int:
    """计算编辑距离"""
    if len(s1) < len(s2):
        return _levenshtein(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev = range(len(s2) + 1)
    for i, c1 in enumerate(s1):
        curr = [i + 1]
        for j, c2 in enumerate(s2):
            insertions = prev[j + 1] + 1
            deletions = curr[j] + 1
            substitutions = prev[j] + (c1 != c2)
            curr.append(min(insertions, deletions, substitutions))
        prev = curr
    return prev[len(s2)]


def _get_pinyin(s: str) -> str:
    """获取中文字符串的拼音（无声调）"""
    try:
        from pypinyin import pinyin, Style
        return "".join(p[0] for p in pinyin(s, style=Style.NORMAL))
    except ImportError:
        return ""


def _find_suspects(name: str, existing: list) -> list:
    """查找疑似重复项，返回 [{name, rule}] 列表"""
    suspects = []
    py_name = _get_pinyin(name)
    for item in existing:
        rules = []
        # 规则1：包含关系
        if name in item or item in name:
            rules.append("包含关系")
        # 规则2：编辑距离<=1
        dist = _levenshtein(name, item)
        if dist == 1:
            rules.append("编辑距离1")
        # 规则3：拼音相同
        if py_name and _get_pinyin(item) == py_name:
            rules.append("拼音相同")
        if rules:
            suspects.append({"name": item, "rule": "+".join(rules)})
    return suspects


# 暂存导入待确认的suspects（内存缓存，按config_type分组）
_pending_imports: dict = {}


@router.post("/config/{config_type}")
async def import_config(config_type: str, file: UploadFile = File(...)):
    """导入配置项（部门/岗位/工种/角色），含重复检测"""
    if config_type not in CONFIG_TYPE_LABELS:
        raise HTTPException(400, f"不支持的配置类型: {config_type}")
    _validate_excel(file)
    try:
        import openpyxl
        import json
        from core.database import get_by
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        raise HTTPException(400, f"Excel 解析失败: {str(e)}")

    label = CONFIG_TYPE_LABELS[config_type]
    row = get_by("system_config", "key", config_type)
    existing = json.loads(row.get("value", "[]")) if row else []

    results = {"total": len(rows), "success": 0, "skipped": 0, "suspects": [], "errors": []}
    to_add = []       # 确认新增的
    suspect_list = []  # 待确认的

    for i, r in enumerate(rows):
        line_no = i + 2
        name = str(r[0]).strip() if r and r[0] else ""
        if not name:
            results["errors"].append(f"第{line_no}行：{label}名称为空")
            continue
        # 精确重复
        if name in existing or name in to_add:
            results["skipped"] += 1
            results["errors"].append(f"第{line_no}行：'{name}'已存在，跳过")
            continue
        # 疑似重复检测
        suspects = _find_suspects(name, existing)
        if suspects:
            suspect_list.append({"name": name, "line": line_no, "similar_to": suspects})
            continue
        to_add.append(name)
        results["success"] += 1

    # 写回数据库（仅确认新增的）
    if to_add:
        _write_config(config_type, existing + to_add, row)
    # 暂存suspects供后续确认
    if suspect_list:
        _pending_imports[config_type] = {"existing_snapshot": existing, "suspects": suspect_list}
        results["suspects"] = suspect_list
    return results


def _write_config(config_type: str, items: list, existing_row=None):
    """写入配置到数据库"""
    import json
    from core.database import get_by, add as db_add
    val = json.dumps(items, ensure_ascii=False)
    if existing_row:
        from core.database import update as db_update
        db_update("system_config", existing_row["id"], {"value": val})
    else:
        from core.utils import generate_id
        db_add("system_config", {"id": generate_id(), "key": config_type, "value": val})


@router.post("/config/{config_type}/confirm")
async def confirm_import_config(config_type: str, body: dict):
    """确认导入suspects中的项
    body: {"confirmed": ["安全课", "生产部车间"], "rejected": ["安全课2"]}
    """
    if config_type not in CONFIG_TYPE_LABELS:
        raise HTTPException(400, f"不支持的配置类型: {config_type}")
    import json
    from core.database import get_by

    confirmed = body.get("confirmed", [])
    row = get_by("system_config", "key", config_type)
    existing = json.loads(row.get("value", "[]")) if row else []

    added = 0
    for name in confirmed:
        name = name.strip()
        if name and name not in existing:
            existing.append(name)
            added += 1

    if added > 0:
        _write_config(config_type, existing, row)

    # 清理暂存
    _pending_imports.pop(config_type, None)

    label = CONFIG_TYPE_LABELS[config_type]
    return {"confirmed": len(confirmed), "added": added, "message": f"已确认添加{added}个{label}"}
